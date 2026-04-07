from fastapi import APIRouter, Body, Response, Cookie, HTTPException, Request, Header
from fastapi.responses import StreamingResponse
from app.database import engine
from app.models import Base
from openpyxl import Workbook
from openpyxl.styles import Font
from app.config import settings
from sqlalchemy import text
from dotenv import load_dotenv
from time import time
from fastapi import Depends
from datetime import datetime, timedelta

import io
import os
import jwt

attempts = {}

router = APIRouter()

# Initialisation BD

@router.get("/init-db")
def init_db():
    Base.metadata.create_all(bind=engine)
    return {"message": "Tables créées"}

# Config

load_dotenv()

PIN_CODE = os.getenv("PIN_CODE")
ADMIN_PIN = os.getenv("ADMIN_PIN")
SECRET_KEY = os.getenv("SECRET_KEY")
if not SECRET_KEY:
    raise Exception("SECRET_KEY non défini !")

# Json Web Token

def create_token(role):
    payload = {
        "role": role,
        "exp": datetime.utcnow() + timedelta(hours=4)
    }
    return jwt.encode(payload, SECRET_KEY, algorithm="HS256")

def verify_token(authorization: str = Header(None)):

    if not authorization:
        raise HTTPException(status_code=403, detail="Token manquant")
    if not authorization.startswith("Bearer "):
        raise HTTPException(status_code=403, detail="Format invalide")
    try:
        token = authorization.replace("Bearer ", "")
        data = jwt.decode(token, SECRET_KEY, algorithms=["HS256"])
        return data.get("role")
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=403, detail="Token expiré")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=403, detail="Token invalide")


# Login Pin = pour voir si Admin ou user

@router.get("/check-access")
def check_access(code: str):

    if code == ADMIN_PIN:
        return {"ok": True, "token": create_token("admin")}
    if code == PIN_CODE:
        return {"ok": True, "token": create_token("user")}
    return {"ok": False}

# Joueurs 

@router.get("/joueurs")
def get_players(request: Request):
    session = request.cookies.get("session")

    if session not in ["user", "admin"]:
        raise HTTPException(status_code=403, detail="Accès interdit")

    with engine.connect() as conn:
        result = conn.execute(text("""
            SELECT id, name, ranking
            FROM players
            ORDER BY ranking DESC
        """))
        joueurs = [
            {"id": row.id, "name": row.name, "ranking": row.ranking}
            for row in result
        ]
    return joueurs


# Jour des Matchs

@router.get("/match-days")
def get_match_days():
    with engine.connect() as conn:
        result = conn.execute(text("""
            SELECT id, code, date, is_home, day_type
            FROM match_days
            ORDER BY id
        """))
        return [dict(row._mapping) for row in result]

# Initialisation 

@router.get("/init-match-days")
def init_match_days():
    with engine.begin() as conn:
        for day in settings.MATCH_DAYS:
            conn.execute(text("""
                INSERT INTO match_days (id, code, date, is_home, day_type)
                VALUES (:id, :code, :date, :is_home, :day_type)
                ON CONFLICT (id) DO UPDATE SET
                    code = EXCLUDED.code,
                    date = EXCLUDED.date,
                    is_home = EXCLUDED.is_home,
                    day_type = EXCLUDED.day_type
            """), day)

    print(" Match days initialisés")
    return {"message": "Journées configurées"}

# Jour a selectionner

@router.get("/init-slots")
def init_slots():
    with engine.begin() as conn:
        for day_id in range(1, 15):
            slots = [
                "samedi_aprem",
                "dimanche_matin",
                "dimanche_aprem"
            ]
            for label in slots:
                conn.execute(text("""
                    INSERT INTO match_slots (match_day_id, label)
                    VALUES (:day_id, :label)
                    ON CONFLICT (match_day_id, label) DO NOTHING
                """), {
                    "day_id": day_id,
                    "label": label
                })

    print("✅ Slots initialisés")
    return {"message": "Slots créés"}

# Pour un future si app dynamique

@router.get("/slots")
def get_slots():
    with engine.connect() as conn:
        result = conn.execute(text("""
            SELECT s.id, d.code, s.label
            FROM match_slots s
            JOIN match_days d ON d.id = s.match_day_id
            ORDER BY d.id, s.id
        """))
        return [dict(row._mapping) for row in result]


# Dispo

@router.get("/dispos/{match_day_id}")
def get_dispos(match_day_id: int):
    try:
        with engine.connect() as conn:
            result = conn.execute(text("""
                SELECT 
                    p.name,
                    p.ranking,
                    STRING_AGG(
                        s.label || ':' || a.availability,
                        ','
                    ) AS slots
                FROM availabilities a
                JOIN players p ON p.id = a.player_id
                JOIN match_slots s ON s.id = a.slot_id
                WHERE s.match_day_id = :day_id
                GROUP BY p.name, p.ranking
                ORDER BY p.ranking DESC
            """), {"day_id": match_day_id})
            return [dict(row._mapping) for row in result]

    except Exception as e:
        print("🔥 ERREUR BACKEND:", e)
        return {"error": str(e)}

# Joueurs

@router.get("/player/{license}")
def get_player(license: str):
    with engine.connect() as conn:

        player = conn.execute(text("""
            SELECT id, name
            FROM players
            WHERE license = :license
        """), {"license": license}).fetchone()

        if not player:
            return {"name": None}

        rows = conn.execute(text("""
            SELECT 
                a.match_day_id,
                s.label,
                a.availability
            FROM availabilities a
            JOIN match_slots s ON s.id = a.slot_id
            WHERE a.player_id = :player_id
        """), {"player_id": player.id}).fetchall()

        availability = {}

        for row in rows:
            day_id = row.match_day_id
            if day_id not in availability:
                availability[day_id] = []
            availability[day_id].append({
                "label": row.label,
                "available": row.availability
            })
        availability_list = [
            {
                "match_day_id": day_id,
                "slots": slots
            }
            for day_id, slots in availability.items()
        ]
        return {
            "name": player.name,
            "availability": availability_list
        }

# Availability

@router.post("/availability")
def add_availability(
    data: dict = Body(...),
    role: str = Depends(verify_token)
):

    print("ROLE:", role)
    print("DATA:", data)
    with engine.begin() as conn:
        player = conn.execute(text("""
            SELECT id FROM players WHERE license = :license
        """), {"license": data["license"]}).fetchone()
        if not player:
            raise HTTPException(status_code=400, detail="Licence invalide")
        player_id = player.id
        slots_db = conn.execute(text("""
            SELECT id, label
            FROM match_slots
            WHERE match_day_id = :day_id
        """), {"day_id": data["match_day_id"]}).fetchall()
        slot_map = {row.label: row.id for row in slots_db}
        print("SLOT MAP:", slot_map)

        for slot in data["slots"]:
            label = slot["label"]
            available = slot["available"]
            if label not in slot_map:
                continue
            slot_id = slot_map[label]
            conn.execute(text("""
                INSERT INTO availabilities (player_id, slot_id, match_day_id, availability)
                VALUES (:player_id, :slot_id, :match_day_id, :availability)
                ON CONFLICT (player_id, slot_id, match_day_id)
                DO UPDATE SET availability = EXCLUDED.availability
            """), {
                "player_id": player_id,
                "slot_id": slot_id,
                "match_day_id": data["match_day_id"],
                "availability": available
            })

    return {"message": "Disponibilités enregistrées"}

# Export Excel

@router.get("/export-excel/{match_day_id}")
def export_excel(match_day_id: int):

    with engine.connect() as conn:
        result = conn.execute(text("""
            SELECT 
                p.name,
                p.ranking,
                s.label
            FROM availabilities a
            JOIN players p ON p.id = a.player_id
            JOIN match_slots s ON s.id = a.slot_id
            WHERE s.match_day_id = :day_id
              AND a.availability = 'disponible'
            ORDER BY s.label, p.ranking DESC
        """), {"day_id": match_day_id})

        rows = result.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = f"J{match_day_id}"
    bold = Font(bold=True)
    ws.cell(row=1, column=1, value="Prénom Nom").font = bold
    ws.cell(row=1, column=2, value="Points").font = bold
    grouped = {}
    for row in rows:
        grouped.setdefault(row.label, []).append(row)

    row_idx = 3
    order = ["samedi_aprem", "dimanche_matin", "dimanche_aprem"]

    for label in order:
        if label not in grouped:
            continue
        ws.cell(row=row_idx, column=1, value=label)
        row_idx += 1
        for p in grouped[label]:
            ws.cell(row=row_idx, column=1, value=p.name)
            ws.cell(row=row_idx, column=2, value=p.ranking)
            row_idx += 1
        row_idx += 1
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=journee_{match_day_id}.xlsx"
        }
    )
